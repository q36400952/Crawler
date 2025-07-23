import os
import requests
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
import time

IMDB_EMAIL = "q36400952@gmail.com"
IMDB_PASSWORD = "gxuouar2"

def get_driver():
    options = Options()
    # options.add_argument("start-maximized")
    # options.add_argument("--headless")
    options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                         "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.0.0 Safari/537.36")
    # options.add_experimental_option("excludeSwitches", ["enable-automation"])
    # options.add_experimental_option("useAutomationExtension", False)

    driver = webdriver.Chrome(options=options)
    driver.execute_cdp_cmd('Page.addScriptToEvaluateOnNewDocument', {'source': 'Object.defineProperty(navigator, "webdriver", {get: () => undefined})'})
    return driver

def login_imdb(driver, email, password):
    driver.get("https://www.imdb.com/what-to-watch/fan-favorites/")

    print("正在登入 IMDb")

    sign_in_method = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "a.imdb-header__signin-text")))
    sign_in_method.click()

    imdb_signin_btn = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.CSS_SELECTOR, 'a[href*="/ap/signin"]')))
    imdb_signin_btn.click()

    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "ap_email")))

    driver.find_element(By.ID, "ap_email").send_keys(email)
    driver.find_element(By.ID, "ap_password").send_keys(password)

    driver.find_element(By.ID, "signInSubmit").click()

    print("完成驗證後，按 Enter 繼續...")
    input()

    print("登入完成！")

def wait_for_movies(driver):
    WebDriverWait(driver, 15).until(
        EC.presence_of_element_located((By.CSS_SELECTOR, "div.ipc-poster-card"))
    )
    time.sleep(3)

def parse_current_page(driver, section_name):
    soup = BeautifulSoup(driver.page_source, "html.parser")
    movies = soup.select("div.ipc-poster-card")

    if not movies:
        print(f"找不到電影：{section_name}，請檢查 debug_{section_name}.html")
        with open(f"debug_{section_name}.html", "w", encoding="utf-8") as f:
            f.write(driver.page_source)
        return []

    print(f"分頁：{section_name}，共 {len(movies)} 部作品")

    img_folder = "temp_images"
    os.makedirs(img_folder, exist_ok=True)

    results = []
    for idx, movie in enumerate(movies):
        title_tag = movie.select_one("span[data-testid='title']")
        rating_tag = movie.select_one("span.ipc-rating-star--rating")
        img_tag = movie.select_one("img")

        title = title_tag.text.strip() if title_tag else "無標題"
        rating = rating_tag.text.strip() if rating_tag else "無評分"
        img_url = img_tag['src'] if img_tag and img_tag.has_attr('src') else ""

        print(f"{title} | 評分: {rating} | 圖片連結: {img_url if img_url else '無圖片'}")

        img_path = ""
        if img_url:
            try:
                img_data = requests.get(img_url, timeout=10).content
                img_path = os.path.join(img_folder, f"{section_name}_{idx}.jpg")
                with open(img_path, "wb") as f:
                    f.write(img_data)
            except Exception as e:
                print(f"圖片下載失敗：{title}，原因：{e}")

        results.append({
            "section": section_name,
            "title": title,
            "rating": rating,
            "image_url": img_url,
            "image_path": img_path
        })
    return results

def fetch_imdb_tabs(driver):
    url = "https://www.imdb.com/what-to-watch/fan-favorites/"
    driver.get(url)

    all_results = []

    print("\n抓取分頁：fan-favorites")
    wait_for_movies(driver)
    all_results.extend(parse_current_page(driver, "FAN FAVORITES"))

    tabs = ["TOP PICKS", "FROM YOUR WATCHLIST", "MOST POPULAR"]

    for tab_name in tabs:
        print(f"\n進入分頁：{tab_name}")
        try:
            tab = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, f"//li[.//span[text()='{tab_name}']]"))
            )
            driver.execute_script("arguments[0].click();", tab)
            wait_for_movies(driver)
            all_results.extend(parse_current_page(driver, tab_name))
        except Exception as e:
            print(f"無法載入 {tab_name} 分頁:", e)

    return all_results

def save_to_excel(data, filename="imdb_movies_with_images.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.title = "IMDb Movies"
    ws.append(["圖片", "分頁", "標題", "評分"])

    for index, item in enumerate(data, start=2):
        ws.cell(row=index, column=2, value=item["section"])
        ws.cell(row=index, column=3, value=item["title"])
        ws.cell(row=index, column=4, value=item["rating"])

        ws.row_dimensions[index].height = 100
        if item.get("image_path") and os.path.exists(item["image_path"]):
            img = ExcelImage(item["image_path"])
            img.width = 80
            img.height = 100
            img.anchor = f"A{index}"
            ws.add_image(img)



    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 10

    wb.save(filename)
    print(f"資料與圖片已存成 Excel：{filename}")

if __name__ == "__main__":
    driver = get_driver()
    try:
        login_imdb(driver, IMDB_EMAIL, IMDB_PASSWORD)
        all_data = fetch_imdb_tabs(driver)
        save_to_excel(all_data)
    finally:
        driver.quit()